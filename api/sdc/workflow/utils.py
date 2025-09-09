import io
import csv
import hashlib
from typing import Tuple, Dict, Any
import chardet
import mimetypes
from openpyxl import load_workbook


def compute_checksum(file_bytes: bytes) -> str:
    return hashlib.sha256(file_bytes).hexdigest()


def detect_mime_from_name(filename: str) -> str:
    mime, _ = mimetypes.guess_type(filename)
    return mime or "application/octet-stream"


def sniff_delimiter(sample: str) -> str:
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        return dialect.delimiter
    except Exception:
        return ","


def preflight(file_bytes: bytes, filename: str) -> Dict[str, Any]:
    mime = detect_mime_from_name(filename)
    checksum = compute_checksum(file_bytes)
    report: Dict[str, Any] = {"mime": mime, "checksum": checksum, "headers": [], "issues": []}
    if filename.lower().endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active
        headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
        report["fileFormat"] = "xlsx"
        report["headers"] = headers
    else:
        # assume text delimited
        detected = chardet.detect(file_bytes[:4096])
        encoding = detected.get("encoding") or "utf-8"
        sample = file_bytes[:2048].decode(encoding=encoding, errors="ignore")
        delimiter = sniff_delimiter(sample)
        report["fileFormat"] = "csv"
        report["encoding"] = encoding
        report["delimiter"] = delimiter
        reader = csv.reader(io.StringIO(file_bytes.decode(encoding, errors="ignore")), delimiter=delimiter)
        try:
            headers = next(reader)
        except StopIteration:
            headers = []
        headers = [h.strip() for h in headers]
        report["headers"] = headers
    # simple header checks
    if len(report["headers"]) != len(set([h.lower() for h in report["headers"] if h])):
        report["issues"].append("duplicate_headers")
    if any(not h for h in report["headers"]):
        report["issues"].append("empty_headers")
    return report
