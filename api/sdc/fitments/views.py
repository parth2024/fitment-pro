from django.shortcuts import render
from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q, Count, Case, When, IntegerField, Min, Max
from django.http import StreamingHttpResponse, HttpResponse, JsonResponse
from django.core.paginator import Paginator
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from .models import Fitment, FitmentUploadSession, FitmentValidationResult, PotentialVehicleConfiguration
from .validators import validate_fitment_row
from tenants.utils import get_tenant_from_request, filter_queryset_by_tenant, get_tenant_id_from_request
import os
import csv
import json
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO
import uuid
from datetime import datetime
import logging
import pandas as pd
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

logger = logging.getLogger(__name__)


# Create your views here.


def _apply_filters(queryset, params):
    """Apply comprehensive filtering to fitments queryset"""
    # Entity filtering - handle entity_ids parameter
    entity_ids = params.get("entity_ids")
    if entity_ids:
        # Split comma-separated entity IDs and filter by tenant
        entity_id_list = [eid.strip() for eid in entity_ids.split(',') if eid.strip()]
        if entity_id_list:
            queryset = queryset.filter(tenant_id__in=entity_id_list)
    
    # Global search
    search = params.get("search")
    if search:
        q = Q(partId__icontains=search) | Q(makeName__icontains=search) | Q(modelName__icontains=search) | Q(fitmentTitle__icontains=search) | Q(fitmentDescription__icontains=search)
        queryset = queryset.filter(q)
    
    # Vehicle filter with match all/any mode
    # 
    # FILTER MODE EXPLANATION:
    # 
    # MATCH_ANY (OR logic): Returns fitments that match ANY of the provided configurations
    #   Example input:
    #     2026
    #     2025|Toyota|Camry
    #   Result: Returns fitments from year 2026 OR (year 2025 AND Toyota AND Camry)
    #   SQL equivalent: WHERE (year=2026) OR (year=2025 AND make='Toyota' AND model='Camry')
    #
    # MATCH_ALL (AND logic): Returns fitments that match ALL of the provided configurations
    #   Example input:
    #     2026
    #     2025|Toyota|Camry
    #   Result: Returns fitments that match BOTH (year=2026) AND (year=2025 AND Toyota AND Camry)
    #   This is usually an impossible condition, so MATCH_ALL is typically used with
    #   overlapping/cumulative criteria rather than different configurations.
    #   SQL equivalent: WHERE (year=2026) AND (year=2025 AND make='Toyota' AND model='Camry')
    #
    vehicle_filter = params.get("vehicleFilter")
    filter_mode = params.get("filterMode", "MATCH_ALL")
    
    if vehicle_filter:
        # Parse vehicle filter format: Year | Make | Model | Submodel | DriveType | FuelType | NumDoors
        # e.g: 2008|Dodge|Ram 1500|*|4WD|*|*
        # Or single value: 2026 (just a year)
        lines = vehicle_filter.strip().split('\n')
        vehicle_conditions = []
        
        for line in lines:
            line = line.strip()
            if not line:
                continue
            
            # Check if it's a single value (e.g., just a year) or pipe-separated format
            if '|' not in line:
                # Single value - treat as year filter
                try:
                    year_value = int(line.strip())
                    vehicle_q = Q(year=year_value)
                    vehicle_conditions.append(vehicle_q)
                    continue
                except ValueError:
                    # If not a valid year, skip this line
                    continue
                
            # Pipe-separated format
            parts = [part.strip() for part in line.split('|')]
            
            # Build Q object for this vehicle configuration
            vehicle_q = Q()
            
            # Handle different number of parts (at least 1, up to 7)
            if len(parts) >= 1 and parts[0] and parts[0] != '*':
                try:
                    vehicle_q &= Q(year=int(parts[0]))
                except ValueError:
                    pass
            
            if len(parts) >= 2 and parts[1] and parts[1] != '*':
                vehicle_q &= Q(makeName__icontains=parts[1])
            
            if len(parts) >= 3 and parts[2] and parts[2] != '*':
                vehicle_q &= Q(modelName__icontains=parts[2])
            
            if len(parts) >= 4 and parts[3] and parts[3] != '*':
                vehicle_q &= Q(subModelName__icontains=parts[3])
            
            if len(parts) >= 5 and parts[4] and parts[4] != '*':
                vehicle_q &= Q(driveTypeName__icontains=parts[4])
            
            if len(parts) >= 6 and parts[5] and parts[5] != '*':
                vehicle_q &= Q(fuelTypeName__icontains=parts[5])
            
            if len(parts) >= 7 and parts[6] and parts[6] != '*':
                vehicle_q &= Q(bodyNumDoors__icontains=parts[6])
            
            # Only add condition if at least one filter was set
            if vehicle_q.children or vehicle_q.connector == Q.AND:
                vehicle_conditions.append(vehicle_q)
        
        if vehicle_conditions:
            if filter_mode == "MATCH_ANY":
                # MATCH_ANY: Return fitments that match ANY of the provided configurations
                # Example: If filter has "2026" and "2025|Toyota|Camry"
                # Result: Returns fitments from year 2026 OR (year 2025 AND Toyota Camry)
                # Logic: (condition1) OR (condition2) OR (condition3) ...
                combined_q = vehicle_conditions[0]
                for condition in vehicle_conditions[1:]:
                    combined_q |= condition
                queryset = queryset.filter(combined_q)
            else:
                # MATCH_ALL: Return fitments that match ALL of the provided configurations
                # Note: This is typically used when you want fitments that satisfy multiple criteria
                # Example: If filter has multiple lines, returns fitments matching ALL conditions
                # Logic: (condition1) AND (condition2) AND (condition3) ...
                # WARNING: For most use cases, MATCH_ANY is more useful than MATCH_ALL
                # because MATCH_ALL with different configurations rarely returns results
                combined_q = vehicle_conditions[0]
                for condition in vehicle_conditions[1:]:
                    combined_q &= condition
                queryset = queryset.filter(combined_q)
    
    # Part IDs filtering
    part_ids = params.get("part_ids")
    if part_ids:
        part_id_list = [pid.strip() for pid in part_ids.split(',') if pid.strip()]
        if part_id_list:
            queryset = queryset.filter(partId__in=part_id_list)
    
    # Column-wise filtering - support both direct and column_ prefixed parameters
    def get_param_value(param_name):
        """Get parameter value, checking both direct and column_ prefixed versions"""
        return params.get(param_name) or params.get(f"column_{param_name}")
    
    if get_param_value("partId"):
        queryset = queryset.filter(partId__icontains=get_param_value("partId"))
    
    if get_param_value("itemStatus"):
        queryset = queryset.filter(itemStatus__icontains=get_param_value("itemStatus"))
    
    if get_param_value("yearFrom"):
        try:
            queryset = queryset.filter(year__gte=int(get_param_value("yearFrom")))
        except ValueError:
            pass
    
    if get_param_value("yearTo"):
        try:
            queryset = queryset.filter(year__lte=int(get_param_value("yearTo")))
        except ValueError:
            pass
    
    if get_param_value("makeName"):
        queryset = queryset.filter(makeName__icontains=get_param_value("makeName"))
    
    if get_param_value("modelName"):
        queryset = queryset.filter(modelName__icontains=get_param_value("modelName"))
    
    if get_param_value("subModelName"):
        queryset = queryset.filter(subModelName__icontains=get_param_value("subModelName"))
    
    if get_param_value("driveTypeName"):
        queryset = queryset.filter(driveTypeName__icontains=get_param_value("driveTypeName"))
    
    if get_param_value("fuelTypeName"):
        queryset = queryset.filter(fuelTypeName__icontains=get_param_value("fuelTypeName"))
    
    if get_param_value("bodyTypeName"):
        queryset = queryset.filter(bodyTypeName__icontains=get_param_value("bodyTypeName"))
    
    if get_param_value("partTypeDescriptor"):
        queryset = queryset.filter(partTypeDescriptor__icontains=get_param_value("partTypeDescriptor"))
    
    if get_param_value("position"):
        queryset = queryset.filter(position__icontains=get_param_value("position"))
    
    if get_param_value("liftHeight"):
        queryset = queryset.filter(liftHeight__icontains=get_param_value("liftHeight"))
    
    if get_param_value("wheelType"):
        queryset = queryset.filter(wheelType__icontains=get_param_value("wheelType"))
    
    if get_param_value("fitmentType"):
        queryset = queryset.filter(fitmentType=get_param_value("fitmentType"))
    
    if get_param_value("createdBy"):
        queryset = queryset.filter(createdBy__icontains=get_param_value("createdBy"))
    
    # Date range filtering
    if get_param_value("createdAtFrom"):
        try:
            date_from = datetime.fromisoformat(get_param_value("createdAtFrom").replace('Z', '+00:00'))
            queryset = queryset.filter(createdAt__gte=date_from)
        except ValueError:
            pass
    
    if get_param_value("createdAtTo"):
        try:
            date_to = datetime.fromisoformat(get_param_value("createdAtTo").replace('Z', '+00:00'))
            queryset = queryset.filter(createdAt__lte=date_to)
        except ValueError:
            pass
    
    if get_param_value("updatedAtFrom"):
        try:
            date_from = datetime.fromisoformat(get_param_value("updatedAtFrom").replace('Z', '+00:00'))
            queryset = queryset.filter(updatedAt__gte=date_from)
        except ValueError:
            pass
    
    if get_param_value("updatedAtTo"):
        try:
            date_to = datetime.fromisoformat(get_param_value("updatedAtTo").replace('Z', '+00:00'))
            queryset = queryset.filter(updatedAt__lte=date_to)
        except ValueError:
            pass
    
    return queryset


def _apply_sort(queryset, sort_by: str | None, sort_order: str | None):
    """Apply sorting to fitments queryset"""
    allowed = {
        "partId", "makeName", "modelName", "year", "updatedAt", "createdAt", 
        "itemStatus", "subModelName", "driveTypeName", "fuelTypeName", 
        "bodyTypeName", "partTypeDescriptor", "position", "liftHeight", 
        "wheelType", "fitmentType", "createdBy", "updatedBy", "quantity"
    }
    field = sort_by if sort_by in allowed else "updatedAt"
    if sort_order == "desc":
        field = f"-{field}"
    return queryset.order_by(field)


@api_view(["GET", "POST", "DELETE"])
def fitments_root(request):
    if request.method == "GET":
        params = request.query_params
        
        # Check if entity_ids parameter is provided
        entity_ids = params.get("entity_ids")
        
        if entity_ids:
            # If entity_ids is provided, start with all fitments and let _apply_filters handle the filtering
            qs = Fitment.objects.all()
        else:
            # Handle tenant filtering - use default tenant if no authenticated user
            try:
                qs = filter_queryset_by_tenant(Fitment.objects.all(), request)
            except:
                # If no authenticated user, get all fitments (for testing)
                # In production, this should be restricted
                qs = Fitment.objects.all()
        
        qs = _apply_filters(qs, params)
        qs = _apply_sort(qs, params.get("sortBy"), params.get("sortOrder"))
        # pagination
        page = int(params.get("page", 1))
        page_size = int(params.get("pageSize", 50))
        start = (page - 1) * page_size
        end = start + page_size
        total = qs.count()
        items = list(qs[start:end].values())
        return Response({
            "fitments": items, 
            "totalCount": total,
            "tenant_id": get_tenant_id_from_request(request) if request.user.is_authenticated else None
        })

    if request.method == "POST":
        payload = request.data or {}
        part_ids = payload.get("partIDs") or []
        confs = payload.get("configurationIDs") or []
        part_id = part_ids[0] if part_ids else ""
        
        # Get current tenant - handle case where user is not authenticated
        try:
            tenant = get_tenant_from_request(request)
        except:
            # If no authenticated user, try to get default tenant
            try:
                from tenants.models import Tenant
                tenant = Tenant.objects.filter(is_default=True).first()
                if not tenant:
                    tenant = Tenant.objects.first()
            except:
                tenant = None
        
        if not tenant:
            return Response({
                "error": "No tenant available. Please create a tenant first."
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Minimal create: insert a placeholder fitment
        fitment = Fitment(
            tenant=tenant,
            partId=part_id, itemStatus="Active", itemStatusCode=0,
            baseVehicleId=str(confs[0]) if confs else "",
            year=payload.get("year", 2025), makeName=payload.get("make", "Acura"),
            modelName=payload.get("model", "ADX"), subModelName=payload.get("submodel", "Advance"),
            driveTypeName=payload.get("driveType", "AWD"), fuelTypeName=payload.get("fuelType", "Gas"),
            bodyNumDoors=int(payload.get("numDoors", 4)), bodyTypeName=payload.get("bodyType", "Crossover"),
            ptid=payload.get("ptid", "PT-22"), partTypeDescriptor=payload.get("partTypeDescriptor", "Brake Pads"),
            uom=payload.get("uom", "Set"), quantity=int(payload.get("quantity", 1)),
            fitmentTitle=payload.get("fitmentTitle", "New Fitment"), fitmentDescription=payload.get("fitmentDescription", ""),
            fitmentNotes=payload.get("fitmentNotes", ""), position=payload.get("position", "Front"),
            positionId=int(payload.get("positionId", 1)), liftHeight=payload.get("liftHeight", "Stock"),
            wheelType=payload.get("wheelType", "Alloy"), createdBy="api", updatedBy="api",
        )
        fitment.save()
        return Response({
            "message": "Fitment created successfully",
            "hash": fitment.hash,
            "tenant_id": str(tenant.id)
        })

    # DELETE (bulk by hashes param) - Soft delete
    hashes = request.query_params.getlist("hashes") or []
    deleted_by = request.data.get('deletedBy', 'api_user')
    deleted_count = 0
    
    # Get current tenant - handle case where user is not authenticated
    try:
        tenant = get_tenant_from_request(request)
    except:
        # If no authenticated user, try to get default tenant
        try:
            from tenants.models import Tenant
            tenant = Tenant.objects.filter(is_default=True).first()
            if not tenant:
                tenant = Tenant.objects.first()
        except:
            tenant = None
    
    if not tenant:
        return Response({
            "error": "No tenant available. Please create a tenant first."
        }, status=status.HTTP_400_BAD_REQUEST)
    
    for fitment_hash in hashes:
        try:
            fitment = Fitment.all_objects.get(hash=fitment_hash, tenant=tenant)
            fitment.soft_delete(deleted_by=deleted_by)
            deleted_count += 1
        except Fitment.DoesNotExist:
            continue
    
    return Response({
        "message": f"Deleted {deleted_count} fitments",
        "tenant_id": str(tenant.id)
    })


@api_view(["GET"]) 
def coverage(request):
    """Enhanced coverage analysis with real VCDB data filtered by tenant or entity_ids"""
    qp = request.query_params
    try:
        yf = int(qp.get("yearFrom", 2010))
    except ValueError:
        yf = 2010
    try:
        yt = int(qp.get("yearTo", 2030))
    except ValueError:
        yt = 2030
    
    # Check if entity_ids parameter is provided
    entity_ids = qp.get("entity_ids")
    
    if entity_ids:
        # If entity_ids is provided, filter by those specific entities
        entity_id_list = [eid.strip() for eid in entity_ids.split(',') if eid.strip()]
        if entity_id_list:
            tenant_fitments = Fitment.objects.filter(tenant_id__in=entity_id_list)
            tenant = None  # No single tenant when using entity_ids
            print(f"DEBUG: Coverage API - Using entity_ids filtering: {entity_id_list}")
        else:
            tenant_fitments = Fitment.objects.all()
            tenant = None
    else:
        # Handle tenant filtering - use default tenant if no authenticated user
        try:
            tenant = get_tenant_from_request(request)
            tenant_fitments = filter_queryset_by_tenant(Fitment.objects.all(), request)
            print(f"DEBUG: Coverage API - Using tenant: {tenant.name} (ID: {tenant.id})")
            print(f"DEBUG: Coverage API - X-Tenant-ID header: {request.headers.get('X-Tenant-ID')}")
        except Exception as e:
            # If no authenticated user, get all fitments (for testing)
            # In production, this should be restricted
            tenant_fitments = Fitment.objects.all()
            tenant = None
            print(f"DEBUG: Coverage API - No tenant found, using all fitments. Error: {str(e)}")
            print(f"DEBUG: Coverage API - X-Tenant-ID header: {request.headers.get('X-Tenant-ID')}")
    
    # Get all unique vehicle configurations from fitments (this represents our VCDB universe)
    # We'll use the fitments table as our source of truth for available configurations
    all_configs = tenant_fitments.filter(
        year__gte=yf, 
        year__lte=yt
    ).values('year', 'makeName', 'modelName', 'subModelName').distinct()
    
    # Build universe of all configurations
    universe_set = set()
    make_to_total = {}
    make_to_models = {}
    
    for config in all_configs:
        key = (config['year'], config['makeName'], config['modelName'])
        universe_set.add(key)
        make_to_total[config['makeName']] = make_to_total.get(config['makeName'], 0) + 1
        make_to_models.setdefault(config['makeName'], set()).add(config['modelName'])

    # Get fitted configurations (those with actual fitments)
    fitted_qs = tenant_fitments.filter(year__gte=yf, year__lte=yt)
    fitted_set = set()
    for f in fitted_qs.values("year", "makeName", "modelName"):
        key = (f["year"], f["makeName"], f["modelName"])
        fitted_set.add(key)

    # Count fitted per make
    make_to_fitted = {}
    for (y, mk, md) in fitted_set:
        if (y, mk, md) in universe_set:
            make_to_fitted[mk] = make_to_fitted.get(mk, 0) + 1

    # Build rows
    rows = []
    for make, total in make_to_total.items():
        fitted = make_to_fitted.get(make, 0)
        coverage_percent = int(round((fitted / total) * 100)) if total else 0
        models = sorted(list(make_to_models.get(make, set())))
        rows.append({
            "make": make,
            "configsCount": total,
            "fittedConfigsCount": fitted,
            "coveragePercent": coverage_percent,
            "models": models,
        })

    # Sorting
    sort_by = qp.get("sortBy", "make")
    sort_order = qp.get("sortOrder", "asc")
    def sort_key(r):
        return r.get(sort_by) if sort_by != "models" else len(r.get("models", []))
    rows.sort(key=sort_key, reverse=(sort_order == "desc"))

    return Response({
        "items": rows, 
        "totalCount": len(rows),
        "tenant_id": str(tenant.id) if tenant else None
    })


@api_view(["GET"])
def detailed_coverage(request):
    """Get detailed coverage by model within a make filtered by tenant or entity_ids"""
    make = request.GET.get('make')
    year_from = request.GET.get('yearFrom', 2010)
    year_to = request.GET.get('yearTo', 2030)
    
    if not make:
        return Response({"error": "Make parameter is required"}, status=400)
    
    try:
        year_from = int(year_from)
        year_to = int(year_to)
    except ValueError:
        return Response({"error": "Invalid year parameters"}, status=400)
    
    # Check if entity_ids parameter is provided
    entity_ids = request.GET.get("entity_ids")
    
    if entity_ids:
        # If entity_ids is provided, filter by those specific entities
        entity_id_list = [eid.strip() for eid in entity_ids.split(',') if eid.strip()]
        if entity_id_list:
            tenant_fitments = Fitment.objects.filter(tenant_id__in=entity_id_list)
            tenant = None  # No single tenant when using entity_ids
            print(f"DEBUG: Detailed Coverage API - Using entity_ids filtering: {entity_id_list}")
        else:
            tenant_fitments = Fitment.objects.all()
            tenant = None
    else:
        # Handle tenant filtering
        try:
            tenant = get_tenant_from_request(request)
            tenant_fitments = filter_queryset_by_tenant(Fitment.objects.all(), request)
            print(f"DEBUG: Detailed Coverage API - Using tenant: {tenant.name} (ID: {tenant.id})")
        except Exception as e:
            # If no authenticated user, get all fitments (for testing)
            tenant_fitments = Fitment.objects.all()
            tenant = None
            print(f"DEBUG: Detailed Coverage API - No tenant found, using all fitments. Error: {str(e)}")
    
    # Get all configurations for this make
    all_configs = tenant_fitments.filter(
        makeName=make,
        year__gte=year_from,
        year__lte=year_to
    ).values('year', 'modelName', 'subModelName').distinct()
    
    # Count total configurations by model
    model_totals = {}
    for config in all_configs:
        model = config['modelName']
        model_totals[model] = model_totals.get(model, 0) + 1
    
    # Count fitted configurations by model
    fitted_configs = tenant_fitments.filter(
        makeName=make,
        year__gte=year_from,
        year__lte=year_to
    ).values('year', 'modelName').distinct()
    
    model_fitted = {}
    for config in fitted_configs:
        model = config['modelName']
        model_fitted[model] = model_fitted.get(model, 0) + 1
    
    # Build detailed coverage data
    coverage_data = []
    for model, total in model_totals.items():
        fitted = model_fitted.get(model, 0)
        coverage_percent = round((fitted / total) * 100, 2) if total > 0 else 0
        
        coverage_data.append({
            'model': model,
            'totalConfigurations': total,
            'fittedConfigurations': fitted,
            'coveragePercent': coverage_percent
        })
    
    # Sort by total configurations descending
    coverage_data.sort(key=lambda x: x['totalConfigurations'], reverse=True)
    
    return Response(coverage_data)


@api_view(["GET"])
def coverage_trends(request):
    """Get coverage trends by year for a specific make filtered by tenant or entity_ids"""
    make = request.GET.get('make')
    
    if not make:
        return Response({"error": "Make parameter is required"}, status=400)
    
    # Check if entity_ids parameter is provided
    entity_ids = request.GET.get("entity_ids")
    
    if entity_ids:
        # If entity_ids is provided, filter by those specific entities
        entity_id_list = [eid.strip() for eid in entity_ids.split(',') if eid.strip()]
        if entity_id_list:
            tenant_fitments = Fitment.objects.filter(tenant_id__in=entity_id_list)
            tenant = None  # No single tenant when using entity_ids
            print(f"DEBUG: Coverage Trends API - Using entity_ids filtering: {entity_id_list}")
        else:
            tenant_fitments = Fitment.objects.all()
            tenant = None
    else:
        # Handle tenant filtering
        try:
            tenant = get_tenant_from_request(request)
            tenant_fitments = filter_queryset_by_tenant(Fitment.objects.all(), request)
            print(f"DEBUG: Coverage Trends API - Using tenant: {tenant.name} (ID: {tenant.id})")
        except Exception as e:
            # If no authenticated user, get all fitments (for testing)
            tenant_fitments = Fitment.objects.all()
            tenant = None
            print(f"DEBUG: Coverage Trends API - No tenant found, using all fitments. Error: {str(e)}")
    
    # Get all configurations by year for this make
    yearly_configs = tenant_fitments.filter(
        makeName=make
    ).values('year').annotate(
        total=Count('id', distinct=True)
    ).order_by('year')
    
    # Get fitted configurations by year
    yearly_fitted = tenant_fitments.filter(
        makeName=make
    ).values('year').annotate(
        fitted=Count('id', distinct=True)
    ).order_by('year')
    
    # Combine data
    trends = []
    configs_dict = {item['year']: item['total'] for item in yearly_configs}
    fitted_dict = {item['year']: item['fitted'] for item in yearly_fitted}
    
    for year in sorted(set(configs_dict.keys()) | set(fitted_dict.keys())):
        total = configs_dict.get(year, 0)
        fitted = fitted_dict.get(year, 0)
        coverage_percent = round((fitted / total) * 100, 2) if total > 0 else 0
        
        trends.append({
            'year': year,
            'totalConfigurations': total,
            'fittedConfigurations': fitted,
            'coveragePercent': coverage_percent
        })
    
    return Response({
        "trends": trends,
        "tenant_id": str(tenant.id) if tenant else None
    })


@api_view(["GET"])
def coverage_gaps(request):
    """Find models with low coverage that need attention filtered by tenant or entity_ids"""
    make = request.GET.get('make')
    year_from = request.GET.get('yearFrom', 2010)
    year_to = request.GET.get('yearTo', 2030)
    min_vehicles = int(request.GET.get('minVehicles', 10))
    max_coverage = float(request.GET.get('maxCoverage', 50.0))
    
    if not make:
        return Response({"error": "Make parameter is required"}, status=400)
    
    try:
        year_from = int(year_from)
        year_to = int(year_to)
    except ValueError:
        return Response({"error": "Invalid year parameters"}, status=400)
    
    # Check if entity_ids parameter is provided
    entity_ids = request.GET.get("entity_ids")
    
    if entity_ids:
        # If entity_ids is provided, filter by those specific entities
        entity_id_list = [eid.strip() for eid in entity_ids.split(',') if eid.strip()]
        if entity_id_list:
            tenant_fitments = Fitment.objects.filter(tenant_id__in=entity_id_list)
            tenant = None  # No single tenant when using entity_ids
            print(f"DEBUG: Coverage Gaps API - Using entity_ids filtering: {entity_id_list}")
        else:
            tenant_fitments = Fitment.objects.all()
            tenant = None
    else:
        # Handle tenant filtering
        try:
            tenant = get_tenant_from_request(request)
            tenant_fitments = filter_queryset_by_tenant(Fitment.objects.all(), request)
            print(f"DEBUG: Coverage Gaps API - Using tenant: {tenant.name} (ID: {tenant.id})")
        except Exception as e:
            # If no authenticated user, get all fitments (for testing)
            tenant_fitments = Fitment.objects.all()
            tenant = None
            print(f"DEBUG: Coverage Gaps API - No tenant found, using all fitments. Error: {str(e)}")
    
    # Get all configurations by model
    model_configs = tenant_fitments.filter(
        makeName=make,
        year__gte=year_from,
        year__lte=year_to
    ).values('modelName').annotate(
        total=Count('id', distinct=True)
    ).filter(total__gte=min_vehicles)
    
    # Get fitted configurations by model
    model_fitted = tenant_fitments.filter(
        makeName=make,
        year__gte=year_from,
        year__lte=year_to
    ).values('modelName').annotate(
        fitted=Count('id', distinct=True)
    )
    
    # Find low coverage models
    low_coverage = []
    fitted_dict = {item['modelName']: item['fitted'] for item in model_fitted}
    
    for config in model_configs:
        model = config['modelName']
        total = config['total']
        fitted = fitted_dict.get(model, 0)
        coverage_percent = (fitted / total * 100) if total > 0 else 0
        
        if coverage_percent < max_coverage:
            low_coverage.append({
                'model': model,
                'totalConfigurations': total,
                'fittedConfigurations': fitted,
                'coveragePercent': round(coverage_percent, 2),
                'gap': total - fitted
            })
    
    # Sort by gap size (largest gaps first)
    low_coverage.sort(key=lambda x: x['gap'], reverse=True)
    
    return Response(low_coverage)


@api_view(["GET"]) 
def property_values(request, property_name: str):
    return Response([])


VALIDATION_CACHE = {"result": None}


@api_view(["POST"]) 
def validate(request):
    uploaded = request.FILES.get("fitments")
    if not uploaded:
        return Response({"message": "fitments file is required"}, status=400)

    max_mb = int(os.getenv("MAX_UPLOAD_MB", "10"))
    payload = uploaded.read()
    if len(payload) > max_mb * 1024 * 1024:
        return Response({"message": f"File too large (> {max_mb} MB)"}, status=413)

    # Allow CSV and basic TSV
    text = payload.decode("utf-8", errors="ignore")
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delimiter = dialect.delimiter
    except Exception:
        delimiter = ","

    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    try:
        headers = next(reader)
    except StopIteration:
        headers = []

    headers = [h.strip() for h in headers]
    allowed = {
        "partId","partTypeId","configurationId","quantity","position","liftHeight","wheelType",
        "wheelDiameter1","tireDiameter1","backspacing1","title","description","notes"
    }
    ignored_columns = [h for h in headers if h not in allowed]

    repaired_rows = {}
    invalid_rows = {}

    def set_repair(ridx, col, val):
        repaired_rows.setdefault(ridx, {})[col] = val

    def set_invalid(ridx, col, msg):
        invalid_rows.setdefault(ridx, {})[col] = msg

    # Process rows
    for idx, row in enumerate(reader, start=1):
        row_map = {headers[i]: row[i].strip() if i < len(row) else "" for i in range(len(headers))}

        part_id = row_map.get("partId", "")
        if part_id and not part_id.startswith("P-"):
            set_invalid(idx, "partId", "Invalid format - must start with P-")

        qty_raw = row_map.get("quantity", "").strip()
        if qty_raw:
            try:
                qty = int(qty_raw)
                if qty <= 0:
                    set_invalid(idx, "quantity", "Must be a positive number")
            except ValueError:
                set_invalid(idx, "quantity", "Must be a positive number")

        # Normalize liftHeight
        lh = row_map.get("liftHeight", "")
        if lh:
            norm_lh = lh.strip()
            if norm_lh.lower() == "stock":
                set_repair(idx, "liftHeight", "Stock")

        # Normalize tireDiameter1 casing
        td = row_map.get("tireDiameter1", "")
        if td:
            up = td.upper()
            if up != td:
                set_repair(idx, "tireDiameter1", up)

        # Derive or normalize wheelDiameter1
        wd = row_map.get("wheelDiameter1", "").strip()
        if not wd and td:
            # extract digits after 'R' pattern (e.g., 255/55R18 -> 18)
            up = td.upper()
            if "R" in up:
                try:
                    val = "".join(ch for ch in up.split("R")[-1] if ch.isdigit())
                    if val:
                        set_repair(idx, "wheelDiameter1", val)
                except Exception:
                    pass
        elif wd:
            try:
                iv = int(float(wd))
                if str(iv) != wd:
                    set_repair(idx, "wheelDiameter1", str(iv))
            except Exception:
                # keep as-is
                pass

    result = {
        "repairedRows": repaired_rows,
        "invalidRows": invalid_rows,
        "ignoredColumns": ignored_columns,
    }
    VALIDATION_CACHE["result"] = result
    return Response(result)


@api_view(["POST"]) 
def submit(request):
    if not VALIDATION_CACHE.get("result"):
        return Response({"message": "No validated data in memory. Please validate first."}, status=400)
    # For MVP, we acknowledge submission without DB writes.
    return Response({"message": "Fitments submitted"})


class Echo:
    def write(self, value):
        return value


@api_view(["GET"]) 
def export_csv(request):
    qs = Fitment.objects.all()
    qs = _apply_filters(qs, request.query_params)
    qs = _apply_sort(qs, request.query_params.get("sortBy"), request.query_params.get("sortOrder"))
    pseudo_buffer = Echo()
    writer = csv.writer(pseudo_buffer)
    headers = [
        "hash","partId","itemStatus","itemStatusCode","baseVehicleId","year","makeName","modelName","subModelName",
        "driveTypeName","fuelTypeName","bodyNumDoors","bodyTypeName","ptid","partTypeDescriptor","uom","quantity",
        "fitmentTitle","fitmentDescription","fitmentNotes","position","positionId","liftHeight","wheelType","createdAt","createdBy","updatedAt","updatedBy"
    ]
    def row_iter():
        yield writer.writerow(headers)
        for f in qs.iterator():
            yield writer.writerow([
                f.hash,f.partId,f.itemStatus,f.itemStatusCode,f.baseVehicleId,f.year,f.makeName,f.modelName,f.subModelName,
                f.driveTypeName,f.fuelTypeName,f.bodyNumDoors,f.bodyTypeName,f.ptid,f.partTypeDescriptor,f.uom,f.quantity,
                f.fitmentTitle,f.fitmentDescription or "",f.fitmentNotes or "",f.position,f.positionId,f.liftHeight,f.wheelType,
                f.createdAt.isoformat(),f.createdBy,f.updatedAt.isoformat(),f.updatedBy
            ])
    response = StreamingHttpResponse(row_iter(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="fitments.csv"'
    return response


@api_view(["GET"]) 
def coverage_export(request):
    # Get coverage data directly instead of calling the view function
    qp = request.query_params
    try:
        yf = int(qp.get("yearFrom", 2010))
    except ValueError:
        yf = 2010
    try:
        yt = int(qp.get("yearTo", 2030))
    except ValueError:
        yt = 2030
    
    # Handle tenant filtering
    try:
        tenant = get_tenant_from_request(request)
        tenant_fitments = filter_queryset_by_tenant(Fitment.objects.all(), request)
        print(f"DEBUG: Coverage Export API - Using tenant: {tenant.name} (ID: {tenant.id})")
    except Exception as e:
        # If no authenticated user, get all fitments (for testing)
        tenant_fitments = Fitment.objects.all()
        tenant = None
        print(f"DEBUG: Coverage Export API - No tenant found, using all fitments. Error: {str(e)}")
    
    # Get all unique vehicle configurations from fitments
    all_configs = tenant_fitments.filter(
        year__gte=yf, 
        year__lte=yt
    ).values('year', 'makeName', 'modelName', 'subModelName').distinct()
    
    # Build universe of all configurations
    universe_set = set()
    make_to_total = {}
    make_to_models = {}
    
    for config in all_configs:
        key = (config['year'], config['makeName'], config['modelName'])
        universe_set.add(key)
        make_to_total[config['makeName']] = make_to_total.get(config['makeName'], 0) + 1
        make_to_models.setdefault(config['makeName'], set()).add(config['modelName'])

    # Get fitted configurations
    fitted_qs = tenant_fitments.filter(year__gte=yf, year__lte=yt)
    fitted_set = set()
    for f in fitted_qs.values("year", "makeName", "modelName"):
        key = (f["year"], f["makeName"], f["modelName"])
        fitted_set.add(key)

    # Count fitted per make
    make_to_fitted = {}
    for (y, mk, md) in fitted_set:
        if (y, mk, md) in universe_set:
            make_to_fitted[mk] = make_to_fitted.get(mk, 0) + 1

    # Build rows
    rows = []
    for make, total in make_to_total.items():
        fitted = make_to_fitted.get(make, 0)
        coverage_percent = int(round((fitted / total) * 100)) if total else 0
        models = sorted(list(make_to_models.get(make, set())))
        rows.append({
            "make": make,
            "configsCount": total,
            "fittedConfigsCount": fitted,
            "coveragePercent": coverage_percent,
            "models": models,
        })

    # Sorting
    sort_by = qp.get("sortBy", "make")
    sort_order = qp.get("sortOrder", "asc")
    def sort_key(r):
        return r.get(sort_by) if sort_by != "models" else len(r.get("models", []))
    rows.sort(key=sort_key, reverse=(sort_order == "desc"))
    
    # Create CSV response
    pseudo_buffer = Echo()
    writer = csv.writer(pseudo_buffer)
    headers = ["make", "configsCount", "fittedConfigsCount", "coveragePercent", "models"]
    def row_iter():
        yield writer.writerow(headers)
        for r in rows:
            yield writer.writerow([
                r["make"], r["configsCount"], r["fittedConfigsCount"], r["coveragePercent"], ", ".join(r.get("models", []))
            ])
    response = StreamingHttpResponse(row_iter(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="coverage.csv"'
    return response


@api_view(["GET"]) 
def export_ai_fitments(request):
    """
    Export AI fitments in JSON, XLS, or CSV export_format.
    Query parameters:
    - export_format: json|xls|csv (default: json)
    - session_id: UUID for session tracking (optional)
    - fitment_ids: comma-separated list of fitment hashes (optional)
    """
    format_type = request.query_params.get('export_format', 'json').lower()
    session_id = request.query_params.get('session_id')
    fitment_ids_param = request.query_params.get('fitment_ids', '')
    
    # Parse fitment IDs
    fitment_ids = []
    if fitment_ids_param:
        fitment_ids = [id.strip() for id in fitment_ids_param.split(',') if id.strip()]
    
    # Get fitments based on parameters
    if fitment_ids:
        # Filter by specific fitment IDs
        qs = Fitment.objects.filter(hash__in=fitment_ids)
    elif session_id:
        # For session-based filtering, we'll use a simple approach
        # In a real implementation, you might have a session model
        # For now, we'll filter by createdBy='api' and recent creation
        qs = Fitment.objects.filter(createdBy='api').order_by('-createdAt')[:100]
    else:
        # Get all fitments
        qs = Fitment.objects.all()
    
    # Apply additional filters if needed
    qs = _apply_filters(qs, request.query_params)
    qs = _apply_sort(qs, request.query_params.get("sortBy"), request.query_params.get("sortOrder"))
    
    # Convert to list of dictionaries
    fitments_data = list(qs.values())
    
    if format_type == 'json':
        return Response({
            'session_id': session_id,
            'fitment_ids': fitment_ids,
            'total_count': len(fitments_data),
            'fitments': fitments_data
        })
    
    elif format_type == 'csv':
        return _export_csv_response(fitments_data, 'ai_fitments')
    
    elif format_type in ['xls', 'xlsx']:
        return _export_xls_response(fitments_data, 'ai_fitments')
    
    else:
        return Response({'error': 'Invalid export_format. Use json, csv, or xlsx'}, status=400)


def _export_csv_response(fitments_data, filename_prefix):
    """Helper function to create CSV export response"""
    pseudo_buffer = Echo()
    writer = csv.writer(pseudo_buffer)
    headers = [
        "hash","partId","itemStatus","itemStatusCode","baseVehicleId","year","makeName","modelName","subModelName",
        "driveTypeName","fuelTypeName","bodyNumDoors","bodyTypeName","ptid","partTypeDescriptor","uom","quantity",
        "fitmentTitle","fitmentDescription","fitmentNotes","position","positionId","liftHeight","wheelType","createdAt","createdBy","updatedAt","updatedBy"
    ]
    
    def row_iter():
        yield writer.writerow(headers)
        for f in fitments_data:
            yield writer.writerow([
                f.get('hash', ''), f.get('partId', ''), f.get('itemStatus', ''), f.get('itemStatusCode', ''),
                f.get('baseVehicleId', ''), f.get('year', ''), f.get('makeName', ''), f.get('modelName', ''),
                f.get('subModelName', ''), f.get('driveTypeName', ''), f.get('fuelTypeName', ''),
                f.get('bodyNumDoors', ''), f.get('bodyTypeName', ''), f.get('ptid', ''),
                f.get('partTypeDescriptor', ''), f.get('uom', ''), f.get('quantity', ''),
                f.get('fitmentTitle', ''), f.get('fitmentDescription', '') or "", f.get('fitmentNotes', '') or "",
                f.get('position', ''), f.get('positionId', ''), f.get('liftHeight', ''), f.get('wheelType', ''),
                f.get('createdAt', ''), f.get('createdBy', ''), f.get('updatedAt', ''), f.get('updatedBy', '')
            ])
    
    response = StreamingHttpResponse(row_iter(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}.csv"'
    return response


def _export_xls_response(fitments_data, filename_prefix):
    """Helper function to create XLS export response"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AI Fitments"
    
    # Define headers
    headers = [
        "Hash", "Part ID", "Item Status", "Item Status Code", "Base Vehicle ID", "Year", "Make Name", "Model Name", "Sub Model Name",
        "Drive Type Name", "Fuel Type Name", "Body Num Doors", "Body Type Name", "PTID", "Part Type Descriptor", "UOM", "Quantity",
        "Fitment Title", "Fitment Description", "Fitment Notes", "Position", "Position ID", "Lift Height", "Wheel Type", "Created At", "Created By", "Updated At", "Updated By"
    ]
    
    # Style for headers
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Write data
    for row, fitment in enumerate(fitments_data, 2):
        ws.cell(row=row, column=1, value=fitment.get('hash', ''))
        ws.cell(row=row, column=2, value=fitment.get('partId', ''))
        ws.cell(row=row, column=3, value=fitment.get('itemStatus', ''))
        ws.cell(row=row, column=4, value=fitment.get('itemStatusCode', ''))
        ws.cell(row=row, column=5, value=fitment.get('baseVehicleId', ''))
        ws.cell(row=row, column=6, value=fitment.get('year', ''))
        ws.cell(row=row, column=7, value=fitment.get('makeName', ''))
        ws.cell(row=row, column=8, value=fitment.get('modelName', ''))
        ws.cell(row=row, column=9, value=fitment.get('subModelName', ''))
        ws.cell(row=row, column=10, value=fitment.get('driveTypeName', ''))
        ws.cell(row=row, column=11, value=fitment.get('fuelTypeName', ''))
        ws.cell(row=row, column=12, value=fitment.get('bodyNumDoors', ''))
        ws.cell(row=row, column=13, value=fitment.get('bodyTypeName', ''))
        ws.cell(row=row, column=14, value=fitment.get('ptid', ''))
        ws.cell(row=row, column=15, value=fitment.get('partTypeDescriptor', ''))
        ws.cell(row=row, column=16, value=fitment.get('uom', ''))
        ws.cell(row=row, column=17, value=fitment.get('quantity', ''))
        ws.cell(row=row, column=18, value=fitment.get('fitmentTitle', ''))
        ws.cell(row=row, column=19, value=fitment.get('fitmentDescription', '') or '')
        ws.cell(row=row, column=20, value=fitment.get('fitmentNotes', '') or '')
        ws.cell(row=row, column=21, value=fitment.get('position', ''))
        ws.cell(row=row, column=22, value=fitment.get('positionId', ''))
        ws.cell(row=row, column=23, value=fitment.get('liftHeight', ''))
        ws.cell(row=row, column=24, value=fitment.get('wheelType', ''))
        ws.cell(row=row, column=25, value=str(fitment.get('createdAt', '')))
        ws.cell(row=row, column=26, value=fitment.get('createdBy', ''))
        ws.cell(row=row, column=27, value=str(fitment.get('updatedAt', '')))
        ws.cell(row=row, column=28, value=fitment.get('updatedBy', ''))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename_prefix}.xlsx"'
    return response


@api_view(["GET"]) 
def ai_fitments_list(request):
    """
    Retrieve AI-generated fitments for a session.
    Query parameters:
    - session_id: UUID for session tracking (optional)
    - limit: Number of results to return (default: 100)
    - offset: Number of results to skip (default: 0)
    """
    session_id = request.query_params.get('session_id')
    limit = int(request.query_params.get('limit', 100))
    offset = int(request.query_params.get('offset', 0))
    
    # Get AI-generated fitments (those created by 'api' user)
    qs = Fitment.objects.filter(createdBy='api')
    
    # Apply session filtering if provided
    if session_id:
        # For now, we'll use a simple approach based on creation time
        # In a real implementation, you might have a session model
        qs = qs.order_by('-createdAt')
    
    # Apply additional filters
    qs = _apply_filters(qs, request.query_params)
    qs = _apply_sort(qs, request.query_params.get("sortBy"), request.query_params.get("sortOrder"))
    
    # Apply pagination
    total_count = qs.count()
    fitments = list(qs[offset:offset + limit].values())
    
    return Response({
        'session_id': session_id,
        'total_count': total_count,
        'limit': limit,
        'offset': offset,
        'fitments': fitments
    })


@api_view(["GET"]) 
def applied_fitments_list(request):
    """
    Get all applied fitments with optional session filter.
    Query parameters:
    - session_id: UUID for session tracking (optional)
    - limit: Number of results to return (default: 100)
    - offset: Number of results to skip (default: 0)
    """
    session_id = request.query_params.get('session_id')
    limit = int(request.query_params.get('limit', 100))
    offset = int(request.query_params.get('offset', 0))
    
    # Get all fitments (applied fitments are those in the main Fitment table)
    qs = Fitment.objects.all()
    
    # Apply session filtering if provided
    if session_id:
        # For session-based filtering, we'll use a simple approach
        # In a real implementation, you might have a session model
        qs = qs.filter(createdBy='api').order_by('-createdAt')
    
    # Apply additional filters
    qs = _apply_filters(qs, request.query_params)
    qs = _apply_sort(qs, request.query_params.get("sortBy"), request.query_params.get("sortOrder"))
    
    # Apply pagination
    total_count = qs.count()
    fitments = list(qs[offset:offset + limit].values())
    
    return Response({
        'session_id': session_id,
        'total_count': total_count,
        'limit': limit,
        'offset': offset,
        'fitments': fitments
    })


@api_view(["GET"])
def fitment_detail(request, fitment_hash):
    """Get detailed information for a specific fitment"""
    try:
        # Allow access to soft deleted fitments for detail view
        fitment = Fitment.all_objects.get(hash=fitment_hash)
        fitment_data = {
            'hash': fitment.hash,
            'partId': fitment.partId,
            'itemStatus': fitment.itemStatus,
            'itemStatusCode': fitment.itemStatusCode,
            'baseVehicleId': fitment.baseVehicleId,
            'year': fitment.year,
            'makeName': fitment.makeName,
            'modelName': fitment.modelName,
            'subModelName': fitment.subModelName,
            'driveTypeName': fitment.driveTypeName,
            'fuelTypeName': fitment.fuelTypeName,
            'bodyNumDoors': fitment.bodyNumDoors,
            'bodyTypeName': fitment.bodyTypeName,
            'ptid': fitment.ptid,
            'partTypeDescriptor': fitment.partTypeDescriptor,
            'uom': fitment.uom,
            'quantity': fitment.quantity,
            'fitmentTitle': fitment.fitmentTitle,
            'fitmentDescription': fitment.fitmentDescription,
            'fitmentNotes': fitment.fitmentNotes,
            'position': fitment.position,
            'positionId': fitment.positionId,
            'liftHeight': fitment.liftHeight,
            'wheelType': fitment.wheelType,
            'fitmentType': fitment.fitmentType,
            'dynamicFields': fitment.dynamicFields or {},  # Include dynamic fields with field config references
            'createdAt': fitment.createdAt.isoformat(),
            'createdBy': fitment.createdBy,
            'updatedAt': fitment.updatedAt.isoformat(),
            'updatedBy': fitment.updatedBy,
        }
        return Response(fitment_data)
    except Fitment.DoesNotExist:
        return Response(
            {"error": "Fitment not found"}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Error getting fitment detail: {str(e)}")
        return Response(
            {"error": "Failed to get fitment details"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(["PUT", "PATCH"])
def update_fitment(request, fitment_hash):
    """Update a specific fitment"""
    try:
        # Allow updating soft deleted fitments (they can be restored during update)
        fitment = Fitment.all_objects.get(hash=fitment_hash)
        
        # Get the data from request
        data = request.data
        
        # Update allowed fields
        allowed_fields = [
            'itemStatus', 'itemStatusCode', 'year', 'makeName', 'modelName', 
            'subModelName', 'driveTypeName', 'fuelTypeName', 'bodyNumDoors', 
            'bodyTypeName', 'ptid', 'partTypeDescriptor', 'uom', 'quantity',
            'fitmentTitle', 'fitmentDescription', 'fitmentNotes', 'position',
            'positionId', 'liftHeight', 'wheelType', 'fitmentType', 'dynamicFields'
        ]
        
        for field in allowed_fields:
            if field in data:
                setattr(fitment, field, data[field])
        
        # Always update the updatedBy and updatedAt fields
        fitment.updatedBy = data.get('updatedBy', 'api_user')
        
        # If the fitment was soft deleted, restore it
        if fitment.isDeleted:
            fitment.restore(restored_by=fitment.updatedBy)
        else:
            fitment.save()
        
        return Response({
            "message": "Fitment updated successfully",
            "hash": fitment.hash,
            "updatedAt": fitment.updatedAt.isoformat()
        })
        
    except Fitment.DoesNotExist:
        return Response(
            {"error": "Fitment not found"}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Error updating fitment: {str(e)}")
        return Response(
            {"error": "Failed to update fitment"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(["GET"])
def fitment_filter_options(request):
    """Get filter options for fitments (unique values for dropdowns)"""
    try:
        # Get unique values for various filter fields
        options = {
            'itemStatus': list(Fitment.objects.values_list('itemStatus', flat=True).distinct().order_by('itemStatus')),
            'makeName': list(Fitment.objects.values_list('makeName', flat=True).distinct().order_by('makeName')),
            'modelName': list(Fitment.objects.values_list('modelName', flat=True).distinct().order_by('modelName')),
            'driveTypeName': list(Fitment.objects.values_list('driveTypeName', flat=True).distinct().order_by('driveTypeName')),
            'fuelTypeName': list(Fitment.objects.values_list('fuelTypeName', flat=True).distinct().order_by('fuelTypeName')),
            'bodyTypeName': list(Fitment.objects.values_list('bodyTypeName', flat=True).distinct().order_by('bodyTypeName')),
            'partTypeDescriptor': list(Fitment.objects.values_list('partTypeDescriptor', flat=True).distinct().order_by('partTypeDescriptor')),
            'position': list(Fitment.objects.values_list('position', flat=True).distinct().order_by('position')),
            'liftHeight': list(Fitment.objects.values_list('liftHeight', flat=True).distinct().order_by('liftHeight')),
            'wheelType': list(Fitment.objects.values_list('wheelType', flat=True).distinct().order_by('wheelType')),
            'fitmentType': list(Fitment.objects.values_list('fitmentType', flat=True).distinct().order_by('fitmentType')),
            'createdBy': list(Fitment.objects.values_list('createdBy', flat=True).distinct().order_by('createdBy')),
            'yearRange': {
                'min': Fitment.objects.aggregate(min_year=Min('year'))['min_year'] or 2000,
                'max': Fitment.objects.aggregate(max_year=Max('year'))['max_year'] or 2030
            },
            'dynamicFields': list(Fitment.objects.values_list('dynamicFields').distinct().order_by('dynamicFields'))
        }
        
        return Response(options)
        
    except Exception as e:
        logger.error(f"Error getting filter options: {str(e)}")
        return Response(
            {"error": "Failed to get filter options"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(["GET"])
def export_fitments_advanced_csv(request):
    """Export fitments with advanced filtering in CSV or XLSX format"""
    try:
        print(request.query_params)
        
        # Get filtered queryset
        qs = Fitment.objects.all()
        qs = _apply_filters(qs, request.query_params)
        qs = _apply_sort(qs, request.query_params.get("sortBy"), request.query_params.get("sortOrder"))
        
        # Convert to list of dictionaries
        fitments_data = list(qs.values())
        
        
        return _export_csv_response(fitments_data, 'fitments_export')
        
            
    except Exception as e:
        logger.error(f"Error exporting fitments: {str(e)}")
        return Response(
            {"error": "Failed to export fitments"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )

@api_view(["GET"])
def export_fitments_advanced_xlsx(request):
    """Export fitments with advanced filtering in CSV or XLSX format"""
    try:
        print(request.query_params)
        
        # Get filtered queryset
        qs = Fitment.objects.all()
        qs = _apply_filters(qs, request.query_params)
        qs = _apply_sort(qs, request.query_params.get("sortBy"), request.query_params.get("sortOrder"))
        
        # Convert to list of dictionaries
        fitments_data = list(qs.values())
        

        return _export_xls_response(fitments_data, 'fitments_export')
        
            
    except Exception as e:
        logger.error(f"Error exporting fitments: {str(e)}")
        return Response(
            {"error": "Failed to export fitments"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(["DELETE"])
def delete_fitment(request, fitment_hash):
    """Soft delete a specific fitment"""
    try:
        fitment = Fitment.all_objects.get(hash=fitment_hash)
        deleted_by = request.data.get('deletedBy', 'api_user')
        fitment.soft_delete(deleted_by=deleted_by)
        
        return Response({
            "message": "Fitment deleted successfully",
            "hash": fitment_hash,
            "deletedAt": fitment.deletedAt.isoformat() if fitment.deletedAt else None
        })
        
    except Fitment.DoesNotExist:
        return Response(
            {"error": "Fitment not found"}, 
            status=status.HTTP_404_NOT_FOUND
        )
    except Exception as e:
        logger.error(f"Error deleting fitment: {str(e)}")
        return Response(
            {"error": "Failed to delete fitment"}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@csrf_exempt
@require_http_methods(["POST"])
def validate_fitments_csv(request):
    """Validate uploaded CSV file"""
    try:
        # Get uploaded file
        if 'fitments' not in request.FILES:
            return JsonResponse({'error': 'No file uploaded'}, status=400)
            
        csv_file = request.FILES['fitments']
        
        # Validate file type
        if not csv_file.name.lower().endswith(('.csv', '.xlsx', '.xls')):
            return JsonResponse({'error': 'Only CSV and Excel files are allowed'}, status=400)
        
        # Validate file size (10MB limit)
        if csv_file.size > 10 * 1024 * 1024:
            return JsonResponse({'error': 'File size must be less than 10MB'}, status=400)
        
        # Handle tenant filtering
        try:
            tenant = get_tenant_from_request(request)
            print(f"DEBUG: Bulk Upload Validation - Using tenant: {tenant.name} (ID: {tenant.id})")
            print(f"DEBUG: Bulk Upload Validation - X-Tenant-ID header: {request.headers.get('X-Tenant-ID')}")
        except Exception as e:
            tenant = None
            print(f"DEBUG: Bulk Upload Validation - No tenant found. Error: {str(e)}")
            print(f"DEBUG: Bulk Upload Validation - X-Tenant-ID header: {request.headers.get('X-Tenant-ID')}")
        
        # Create session
        session_id = uuid.uuid4()
        session = FitmentUploadSession.objects.create(
            user=request.user if hasattr(request, 'user') and request.user.is_authenticated else None,
            tenant=tenant,  # Add tenant to session
            session_id=session_id,
            status='validating',
            file_name=csv_file.name
        )
        
        # Read CSV/Excel file
        try:
            if csv_file.name.lower().endswith('.csv'):
                df = pd.read_csv(csv_file)
            else:
                df = pd.read_excel(csv_file)
        except Exception as e:
            session.status = 'failed'
            session.save()
            return JsonResponse({'error': f'Error reading file: {str(e)}'}, status=400)
        
        # Update session with total rows
        session.total_rows = len(df)
        session.save()
        
        validation_results = []
        repaired_rows = {}
        invalid_rows = {}
        ignored_columns = []
        
        # Validate each row
        for index, row in df.iterrows():
            row_number = index + 2  # +2 for header and 0-based indexing
            
            validation_result = validate_fitment_row(row, row_number)
            
            if validation_result['is_valid']:
                # Row is valid - store all fields as valid
                for column, value in row.items():
                    FitmentValidationResult.objects.create(
                        session=session,
                        row_number=row_number,
                        column_name=column,
                        original_value=str(value) if not pd.isna(value) else '',
                        is_valid=True
                    )
            elif validation_result['can_repair']:
                # Row can be auto-repaired
                repaired_rows[row_number] = validation_result['repairs']
                for column, value in row.items():
                    corrected_value = validation_result['repairs'].get(column, value)
                    FitmentValidationResult.objects.create(
                        session=session,
                        row_number=row_number,
                        column_name=column,
                        original_value=str(value) if not pd.isna(value) else '',
                        corrected_value=str(corrected_value) if not pd.isna(corrected_value) else '',
                        is_valid=True,
                        error_message='Auto-corrected' if column in validation_result['repairs'] else None
                    )
            else:
                # Row has errors that cannot be auto-repaired
                invalid_rows[row_number] = validation_result['errors']
                for column, value in row.items():
                    error_message = validation_result['errors'].get(column)
                    FitmentValidationResult.objects.create(
                        session=session,
                        row_number=row_number,
                        column_name=column,
                        original_value=str(value) if not pd.isna(value) else '',
                        is_valid=column not in validation_result['errors'],
                        error_message=error_message
                    )
        
        # Calculate statistics
        valid_rows = session.total_rows - len(invalid_rows)
        session.valid_rows = valid_rows
        session.invalid_rows = len(invalid_rows)
        session.status = 'validated'
        session.save()
        
        return JsonResponse({
            'session_id': str(session_id),
            'repairedRows': repaired_rows,
            'invalidRows': invalid_rows,
            'ignoredColumns': ignored_columns,
            'totalRows': session.total_rows,
            'validRows': valid_rows,
            'invalidRowsCount': len(invalid_rows),
            'tenant_id': str(tenant.id) if tenant else None,
            'tenant_name': tenant.name if tenant else None
        })
        
    except Exception as e:
        logger.error(f"Error validating CSV: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@csrf_exempt
@require_http_methods(["POST"])
def submit_validated_fitments(request, session_id):
    """Submit validated fitments to database"""
    try:
        # Handle tenant filtering
        try:
            tenant = get_tenant_from_request(request)
            print(f"DEBUG: Bulk Upload Submit - Using tenant: {tenant.name} (ID: {tenant.id})")
            print(f"DEBUG: Bulk Upload Submit - X-Tenant-ID header: {request.headers.get('X-Tenant-ID')}")
        except Exception as e:
            tenant = None
            print(f"DEBUG: Bulk Upload Submit - No tenant found. Error: {str(e)}")
            print(f"DEBUG: Bulk Upload Submit - X-Tenant-ID header: {request.headers.get('X-Tenant-ID')}")
        
        session = FitmentUploadSession.objects.get(
            session_id=session_id,
            user=request.user if hasattr(request, 'user') and request.user.is_authenticated else None,
            status='validated'
        )
        
        # Get all valid validation results
        valid_results = FitmentValidationResult.objects.filter(
            session=session,
            is_valid=True
        )
        
        if not valid_results.exists():
            return JsonResponse({
                'error': 'No valid fitments to submit'
            }, status=400)
        
        # Group by row number and create fitment records
        rows_data = {}
        for result in valid_results:
            row_num = result.row_number
            if row_num not in rows_data:
                rows_data[row_num] = {}
            
            # Use corrected value if available, otherwise original
            value = result.corrected_value or result.original_value
            rows_data[row_num][result.column_name] = value
        
        # Create fitment records
        created_count = 0
        skipped_count = 0
        errors = []
        
        for row_number, row_data in rows_data.items():
            try:
                # Validate required fields
                required_fields = ['PartID', 'YearID', 'MakeName', 'ModelName', 'PTID']
                for field in required_fields:
                    if not row_data.get(field):
                        raise ValueError(f'Missing required field: {field}')
                
                # Check for existing fitment with same key data (filtered by tenant)
                existing_fitment = Fitment.objects.filter(
                    partId=row_data['PartID'],
                    year=int(row_data['YearID']),
                    makeName=row_data['MakeName'],
                    modelName=row_data['ModelName'],
                    ptid=row_data['PTID'],
                    isDeleted=False
                )
                
                # Apply tenant filtering if tenant exists
                if tenant:
                    existing_fitment = existing_fitment.filter(tenant=tenant)
                
                existing_fitment = existing_fitment.first()
                
                if existing_fitment:
                    skipped_count += 1
                    continue  # Skip creating duplicate
                
                # Create fitment record
                fitment = Fitment.objects.create(
                    partId=row_data['PartID'],
                    year=int(row_data['YearID']),
                    makeName=row_data['MakeName'],
                    modelName=row_data['ModelName'],
                    subModelName=row_data.get('SubModelName', ''),
                    driveTypeName=row_data.get('DriveTypeName', ''),
                    fuelTypeName=row_data.get('FuelTypeName', ''),
                    bodyNumDoors=int(row_data.get('BodyNumDoors', 4)),
                    bodyTypeName=row_data.get('BodyTypeName', ''),
                    tenant=tenant,  # Add tenant to fitment
                    ptid=row_data['PTID'],
                    partTypeDescriptor=row_data.get('PTID', ''),  # Using PTID as descriptor
                    quantity=int(row_data.get('Quantity', 1)),
                    fitmentTitle=row_data.get('FitmentTitle', ''),
                    fitmentDescription=row_data.get('FitmentDescription', ''),
                    fitmentNotes=row_data.get('FitmentNotes', ''),
                    position=row_data.get('Position', ''),
                    positionId=int(row_data.get('PositionId', 1)),
                    liftHeight=row_data.get('LiftHeight', ''),
                    uom=row_data.get('UOM', 'EA'),
                    wheelType=row_data.get('WheelType', ''),
                    fitmentType='manual_fitment',
                    createdBy='bulk_upload',
                    updatedBy='bulk_upload'
                )
                created_count += 1
                
            except Exception as e:
                errors.append(f'Row {row_number}: {str(e)}')
                continue
        
        # Update session status
        session.status = 'submitted'
        session.save()
        
        return JsonResponse({
            'success': True,
            'created_count': created_count,
            'skipped_count': skipped_count,
            'total_rows': len(rows_data),
            'errors': errors,
            'tenant_id': str(tenant.id) if tenant else None,
            'tenant_name': tenant.name if tenant else None,
            'message': f'Successfully created {created_count} fitments for {tenant.name if tenant else "default tenant"}' + (f', skipped {skipped_count} duplicates' if skipped_count > 0 else '')
        })
        
    except FitmentUploadSession.DoesNotExist:
        return JsonResponse({'error': 'Session not found'}, status=404)
    except Exception as e:
        logger.error(f"Error submitting fitments: {str(e)}")
        return JsonResponse({'error': str(e)}, status=500)


@require_http_methods(["GET"])
def get_validation_results(request, session_id):
    """Get validation results for a session"""
    try:
        session = FitmentUploadSession.objects.get(
            session_id=session_id,
            user=request.user if hasattr(request, 'user') and request.user.is_authenticated else None
        )
        
        # Get validation results grouped by row
        validation_results = FitmentValidationResult.objects.filter(session=session)
        
        # Group by row number
        rows_data = {}
        for result in validation_results:
            row_num = result.row_number
            if row_num not in rows_data:
                rows_data[row_num] = {
                    'is_valid': result.is_valid,
                    'errors': {},
                    'repairs': {}
                }
            
            if result.is_valid and result.corrected_value:
                rows_data[row_num]['repairs'][result.column_name] = result.corrected_value
            elif not result.is_valid:
                rows_data[row_num]['errors'][result.column_name] = result.error_message
        
        # Separate valid, invalid, and repaired rows
        invalid_rows = {k: v['errors'] for k, v in rows_data.items() if v['errors']}
        repaired_rows = {k: v['repairs'] for k, v in rows_data.items() if v['repairs']}
        
        return JsonResponse({
            'session_id': str(session_id),
            'invalidRows': invalid_rows,
            'repairedRows': repaired_rows,
            'ignoredColumns': [],  # Could be stored in session
            'totalRows': session.total_rows,
            'validRows': session.valid_rows,
            'invalidRowsCount': session.invalid_rows
        })
        
    except FitmentUploadSession.DoesNotExist:
        return JsonResponse({'error': 'Session not found'}, status=404)
# =============================================================================
# POTENTIAL FITMENTS API (MFT V1)
# =============================================================================

from rest_framework.decorators import api_view
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Q
from .models import Fitment, PotentialVehicleConfiguration
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
import logging

logger = logging.getLogger(__name__)


@api_view(['GET'])
def get_potential_fitments(request, part_id):
    """
    GET /api/fitments/potential/{part_id}/?method=similarity|base-vehicle
    
    Returns potential vehicle configurations for a given part using AI recommendations.
    """
    method = request.GET.get('method', 'similarity')
    
    if method not in ['similarity', 'base-vehicle']:
        return Response(
            {'error': 'Invalid method. Use similarity or base-vehicle'}, 
            status=status.HTTP_400_BAD_REQUEST
        )
    
    try:
        if method == 'similarity':
            recommendations = get_similarity_recommendations(part_id)
        else:
            recommendations = get_base_vehicle_recommendations(part_id)
        
        return Response(recommendations)
    
    except Exception as e:
        logger.error(f"Error in get_potential_fitments: {str(e)}")
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def get_similarity_recommendations(part_id):
    """
    AI-based similarity recommendations using machine learning
    """
    try:
        # Get existing fitments for this part
        existing_fitments = Fitment.objects.filter(partId=part_id, isDeleted=False)
        
        if not existing_fitments.exists():
            return []
        
        # Get all unique vehicle configurations from existing fitments
        existing_configs = existing_fitments.values(
            'baseVehicleId', 'year', 'makeName', 'modelName', 'subModelName',
            'driveTypeName', 'fuelTypeName', 'bodyNumDoors', 'bodyTypeName'
        ).distinct()
        
        if not existing_configs:
            return []
        
        # Get all vehicle configurations from VCDB (we'll use fitments as proxy)
        all_configs = Fitment.objects.values(
            'baseVehicleId', 'year', 'makeName', 'modelName', 'subModelName',
            'driveTypeName', 'fuelTypeName', 'bodyNumDoors', 'bodyTypeName'
        ).distinct()
        
        # Build feature vectors for similarity comparison
        existing_features = []
        all_features = []
        all_configs_list = list(all_configs)
        
        for config in existing_configs:
            feature_vector = [
                config['year'],
                hash(config['makeName']) % 1000,  # Simple encoding
                hash(config['modelName']) % 1000,
                hash(config['subModelName']) % 1000,
                hash(config['driveTypeName']) % 100,
                hash(config['fuelTypeName']) % 100,
                config['bodyNumDoors'],
                hash(config['bodyTypeName']) % 100
            ]
            existing_features.append(feature_vector)
        
        for config in all_configs_list:
            feature_vector = [
                config['year'],
                hash(config['makeName']) % 1000,
                hash(config['modelName']) % 1000,
                hash(config['subModelName']) % 1000,
                hash(config['driveTypeName']) % 100,
                hash(config['fuelTypeName']) % 100,
                config['bodyNumDoors'],
                hash(config['bodyTypeName']) % 100
            ]
            all_features.append(feature_vector)
        
        existing_features = np.array(existing_features)
        all_features = np.array(all_features)
        
        # Calculate similarity using cosine similarity
        similarity_matrix = cosine_similarity(all_features, existing_features)
        
        # Calculate average similarity scores for each configuration
        similarity_scores = {}
        existing_config_keys = set()
        
        for config in existing_configs:
            key = f"{config['year']}_{config['makeName']}_{config['modelName']}_{config['subModelName']}"
            existing_config_keys.add(key)
        
        for i, config in enumerate(all_configs_list):
            config_key = f"{config['year']}_{config['makeName']}_{config['modelName']}_{config['subModelName']}"
            
            if config_key not in existing_config_keys:  # Don't recommend existing ones
                avg_similarity = np.mean(similarity_matrix[i])
                similarity_scores[config_key] = {
                    'score': avg_similarity,
                    'config': config
                }
        
        # Get top recommendations
        top_configs = sorted(similarity_scores.items(), key=lambda x: x[1]['score'], reverse=True)[:50]
        
        # Build response
        recommendations = []
        for config_key, data in top_configs:
            config = data['config']
            score = data['score']
            relevance = max(0, min(100, int(score * 100)))  # Convert to percentage and clamp
            
            # Generate AI explanation based on similarity analysis
            explanation = _generate_similarity_explanation(config, score, existing_configs)
            
            recommendations.append({
                'id': config_key,
                'vehicleId': f"{config['year']}_{config['makeName']}_{config['modelName']}",
                'baseVehicleId': config['baseVehicleId'],
                'year': config['year'],
                'make': config['makeName'],
                'model': config['modelName'],
                'submodel': config['subModelName'],
                'driveType': config['driveTypeName'],
                'fuelType': config['fuelTypeName'],
                'numDoors': config['bodyNumDoors'],
                'bodyType': config['bodyTypeName'],
                'relevance': relevance,
                'method': 'similarity',
                'explanation': explanation
            })
        
        return recommendations
        
    except Exception as e:
        logger.error(f"Error in get_similarity_recommendations: {str(e)}")
        return []


def get_base_vehicle_recommendations(part_id):
    """
    Base vehicle relationship recommendations
    """
    try:
        # Get existing fitments
        existing_fitments = Fitment.objects.filter(partId=part_id, isDeleted=False)
        
        if not existing_fitments.exists():
            return []
        
        # Get base vehicle IDs from existing fitments
        existing_base_vehicle_ids = existing_fitments.values_list('baseVehicleId', flat=True).distinct()
        
        # Find configurations with same base vehicle but different submodels
        recommendations = []
        for base_vehicle_id in existing_base_vehicle_ids:
            # Get all fitments with this base vehicle ID
            related_fitments = Fitment.objects.filter(
                baseVehicleId=base_vehicle_id,
                isDeleted=False
            ).exclude(
                partId=part_id  # Exclude current part
            ).values(
                'baseVehicleId', 'year', 'makeName', 'modelName', 'subModelName',
                'driveTypeName', 'fuelTypeName', 'bodyNumDoors', 'bodyTypeName'
            ).distinct()
            
            for config in related_fitments:
                config_key = f"{config['year']}_{config['makeName']}_{config['modelName']}_{config['subModelName']}"
                
                # Generate AI explanation based on base vehicle relationship
                explanation = _generate_base_vehicle_explanation(config, base_vehicle_id)
                
                recommendations.append({
                    'id': config_key,
                    'vehicleId': f"{config['year']}_{config['makeName']}_{config['modelName']}",
                    'baseVehicleId': config['baseVehicleId'],
                    'year': config['year'],
                    'make': config['makeName'],
                    'model': config['modelName'],
                    'submodel': config['subModelName'],
                    'driveType': config['driveTypeName'],
                    'fuelType': config['fuelTypeName'],
                    'numDoors': config['bodyNumDoors'],
                    'bodyType': config['bodyTypeName'],
                    'relevance': 85,  # High confidence for base vehicle method
                    'method': 'base-vehicle',
                    'explanation': explanation
                })
        
        # Remove duplicates and sort by relevance
        unique_recommendations = {}
        for rec in recommendations:
            if rec['id'] not in unique_recommendations:
                unique_recommendations[rec['id']] = rec
        
        return sorted(unique_recommendations.values(), key=lambda x: x['relevance'], reverse=True)[:50]
        
    except Exception as e:
        logger.error(f"Error in get_base_vehicle_recommendations: {str(e)}")
        return []


@api_view(['GET'])
def get_parts_with_fitments(request):
    """
    GET /api/fitments/parts-with-fitments/
    
    Returns list of parts that have existing fitments.
    """
    try:
        # Get parts that have at least one fitment
        parts_with_fitments = Fitment.objects.filter(isDeleted=False).values(
            'partId'
        ).distinct()
        
        parts_data = []
        for part_info in parts_with_fitments:
            part_id = part_info['partId']
            
            # Get fitment count for this part
            fitment_count = Fitment.objects.filter(partId=part_id, isDeleted=False).count()
            
            # Get first fitment to get part details
            first_fitment = Fitment.objects.filter(partId=part_id, isDeleted=False).first()
            
            if first_fitment:
                parts_data.append({
                    'id': part_id,
                    'description': f"Part {part_id}",  # Could be enhanced with actual part descriptions
                    'unitOfMeasure': first_fitment.uom,
                    'itemStatus': 'Active' if first_fitment.itemStatusCode == 0 else 'Inactive',
                    'fitmentCount': fitment_count
                })
        
        return Response(parts_data)
        
    except Exception as e:
        logger.error(f"Error in get_parts_with_fitments: {str(e)}")
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
def apply_potential_fitments(request):
    """
    POST /api/fitments/apply-potential-fitments/
    
    Creates fitments for selected potential configurations.
    """
    try:
        data = request.data
        part_id = data.get('partId')
        configuration_ids = data.get('configurationIds', [])
        title = data.get('title', 'AI Recommended Fitment')
        description = data.get('description', 'Recommended based on AI analysis')
        quantity = data.get('quantity', 1)
        
        if not part_id or not configuration_ids:
            return Response(
                {'error': 'partId and configurationIds are required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get current tenant from request headers
        try:
            tenant = get_tenant_from_request(request)
            logger.info(f"Using tenant: {tenant.name} (ID: {tenant.id}) for potential fitments")
        except Exception as e:
            logger.error(f"Failed to get tenant from request: {str(e)}")
            return Response(
                {'error': 'No tenant available. Please ensure X-Tenant-ID header is provided.'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        created_fitments = []
        
        for config_id in configuration_ids:
            # Parse configuration ID to get vehicle details
            # Format: year_make_model_submodel
            try:
                parts = config_id.split('_')
                if len(parts) >= 4:
                    year = int(parts[0])
                    make = parts[1]
                    model = parts[2]
                    submodel = parts[3]
                    
                    # Find a matching fitment to get base vehicle ID and other details
                    # Filter by tenant to ensure we get the right reference data
                    matching_fitment = Fitment.objects.filter(
                        year=year,
                        makeName=make,
                        modelName=model,
                        subModelName=submodel,
                        isDeleted=False,
                        tenant=tenant  # Filter by current tenant
                    ).first()
                    
                    # If no exact match found, try to find any fitment with same make/model for reference
                    if not matching_fitment:
                        matching_fitment = Fitment.objects.filter(
                            makeName=make,
                            modelName=model,
                            isDeleted=False,
                            tenant=tenant
                        ).first()
                    
                    # Create new fitment with tenant association
                    # Use matching fitment data if available, otherwise use sensible defaults
                    new_fitment = Fitment.objects.create(
                        tenant=tenant,  # Associate with current tenant
                        partId=part_id,
                        itemStatus='Active',
                        itemStatusCode=0,
                        baseVehicleId=matching_fitment.baseVehicleId if matching_fitment else f"BV-{year}-{make.upper()}-{model.upper()}",
                        year=year,
                        makeName=make,
                        modelName=model,
                        subModelName=submodel,
                        driveTypeName=matching_fitment.driveTypeName if matching_fitment else "AWD",
                        fuelTypeName=matching_fitment.fuelTypeName if matching_fitment else "Gas",
                        bodyNumDoors=matching_fitment.bodyNumDoors if matching_fitment else 4,
                        bodyTypeName=matching_fitment.bodyTypeName if matching_fitment else "Sedan",
                        ptid=matching_fitment.ptid if matching_fitment else "PT-22",
                        partTypeDescriptor=matching_fitment.partTypeDescriptor if matching_fitment else "Brake Pads",
                        uom=matching_fitment.uom if matching_fitment else "Set",
                        quantity=quantity,
                        fitmentTitle=title,
                        fitmentDescription=description,
                        fitmentNotes=f"Created from potential fitment recommendation for {config_id}",
                        position=matching_fitment.position if matching_fitment else "Front",
                        positionId=matching_fitment.positionId if matching_fitment else 1,
                        liftHeight=matching_fitment.liftHeight if matching_fitment else "Stock",
                        wheelType=matching_fitment.wheelType if matching_fitment else "Alloy",
                        fitmentType='potential_fitment',
                        createdBy='ai_system',
                        updatedBy='ai_system'
                    )
                    
                    created_fitments.append({
                        'id': new_fitment.hash,
                        'configId': config_id,
                        'status': 'created',
                        'note': 'Used default values' if not matching_fitment else 'Used matching fitment data'
                    })
                else:
                    created_fitments.append({
                        'id': None,
                        'configId': config_id,
                        'status': 'failed',
                        'error': 'Invalid configuration ID format'
                    })
                    
            except Exception as e:
                created_fitments.append({
                    'id': None,
                    'configId': config_id,
                    'status': 'failed',
                    'error': str(e)
                })
        
        success_count = len([f for f in created_fitments if f['status'] == 'created'])
        failed_count = len([f for f in created_fitments if f['status'] == 'failed'])
        
        return Response({
            'message': f'Successfully created {success_count} fitments. {failed_count} failed.',
            'created': success_count,
            'failed': failed_count,
            'details': created_fitments,
            'tenant_id': str(tenant.id),
            'tenant_name': tenant.name
        })
        
    except Exception as e:
        logger.error(f"Error in apply_potential_fitments: {str(e)}")
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['GET'])
def analytics_dashboard(request):
    """
    GET /api/analytics/dashboard/
    
    Returns aggregated analytics data for the dashboard filtered by current tenant or entity_ids.
    """
    try:
        from datetime import timedelta
        from django.db.models import Count, Q
        from django.utils import timezone
        from tenants.utils import filter_queryset_by_tenant, get_tenant_from_request
        
        # Check if entity_ids parameter is provided
        entity_ids = request.query_params.get("entity_ids")
        
        if entity_ids:
            # If entity_ids is provided, filter by those specific entities
            entity_id_list = [eid.strip() for eid in entity_ids.split(',') if eid.strip()]
            if entity_id_list:
                tenant_fitments = Fitment.objects.filter(tenant_id__in=entity_id_list)
                tenant = None  # No single tenant when using entity_ids
                print(f"DEBUG: Using entity_ids filtering: {entity_id_list}")
            else:
                tenant_fitments = Fitment.objects.all()
                tenant = None
        else:
            # Get tenant-filtered fitments queryset
            try:
                tenant = get_tenant_from_request(request)
                tenant_fitments = filter_queryset_by_tenant(Fitment.objects.all(), request)
                print(f"DEBUG: Tenant found: {tenant.name} (ID: {tenant.id})")
                print(f"DEBUG: X-Tenant-ID header: {request.headers.get('X-Tenant-ID')}")
            except Exception as e:
                # Fallback to all fitments if no tenant found (for testing)
                tenant_fitments = Fitment.objects.all()
                tenant = None
                print(f"DEBUG: No tenant found, using all fitments. Error: {str(e)}")
                print(f"DEBUG: X-Tenant-ID header: {request.headers.get('X-Tenant-ID')}")
        
        print(f"DEBUG: Total fitments before filtering: {Fitment.objects.count()}")
        print(f"DEBUG: Tenant-filtered fitments: {tenant_fitments.count()}")
        
        # Debug: Check fitments with and without tenant
        if tenant:
            fitments_with_tenant = Fitment.objects.filter(tenant_id=tenant.id).count()
            fitments_without_tenant = Fitment.objects.filter(tenant_id__isnull=True).count()
            print(f"DEBUG: Fitments with tenant {tenant.id}: {fitments_with_tenant}")
            print(f"DEBUG: Fitments without tenant (NULL): {fitments_without_tenant}")
            
            # Debug: Show first few fitments and their tenant IDs
            sample_fitments = Fitment.objects.all()[:5]
            for fitment in sample_fitments:
                print(f"DEBUG: Fitment {fitment.hash}: tenant_id = {fitment.tenant_id}")
        # Get total fitments count (tenant-filtered)
        total_fitments = tenant_fitments.count()
        
        # Get manual fitments count (fitmentType = 'manual_fitment')
        manual_fitments = tenant_fitments.filter(fitmentType='manual_fitment').count()
        
        # Get AI fitments count (fitmentType = 'ai_fitment' or 'potential_fitment')
        ai_fitments = tenant_fitments.filter(
            Q(fitmentType='ai_fitment') | Q(fitmentType='potential_fitment')
        ).count()
        
        # Get total parts count (unique partIds)
        total_parts = tenant_fitments.values('partId').distinct().count()
        
        # Get total VCDB configurations count (unique vehicle configurations)
        total_vcdb_configs = tenant_fitments.values(
            'year', 'makeName', 'modelName', 'subModelName'
        ).distinct().count()
        
        # Get recent activity (fitments created in last 30 days)
        thirty_days_ago = timezone.now() - timedelta(days=30)
        recent_activity = tenant_fitments.filter(
            createdAt__gte=thirty_days_ago
        ).count()
        
        # Get fitments by status
        active_fitments = tenant_fitments.filter(itemStatus='Active').count()
        inactive_fitments = tenant_fitments.filter(itemStatus='Inactive').count()
        
        # Get pending review count (fitments with status 'pending' or 'review')
        pending_review_count = tenant_fitments.filter(
            Q(itemStatus__iexact='pending') | Q(itemStatus__iexact='review')
        ).count()
        
        # Get fitments by make (top 5)
        top_makes = tenant_fitments.values('makeName').annotate(
            count=Count('hash')
        ).order_by('-count')[:5]
        
        # Get fitments by year (last 5 years)
        current_year = timezone.now().year
        yearly_stats = []
        for year in range(current_year - 4, current_year + 1):
            year_count = tenant_fitments.filter(year=year).count()
            yearly_stats.append({
                'year': year,
                'count': year_count
            })
        
        # Calculate success rate (active vs total)
        success_rate = round((active_fitments / total_fitments * 100), 1) if total_fitments > 0 else 0
        
        # Calculate coverage percentage (fitments with vehicles vs total possible)
        # This is a simplified calculation - in reality you'd compare against actual VCDB
        coverage_percentage = min(100, round((total_fitments / max(total_vcdb_configs, 1)) * 100), 1)
        
        analytics_data = {
            'totalFitments': total_fitments,
            'manualFitments': manual_fitments,
            'aiFitments': ai_fitments,
            'totalParts': total_parts,
            'totalVcdbConfigs': total_vcdb_configs,
            'recentActivity': recent_activity,
            'activeFitments': active_fitments,
            'inactiveFitments': inactive_fitments,
            'successRate': success_rate,
            'coveragePercentage': coverage_percentage,
            'pendingReviewCount': pending_review_count,
            'topMakes': list(top_makes),
            'yearlyStats': yearly_stats,
            'lastUpdated': timezone.now().isoformat(),
            'tenant': {
                'id': str(tenant.id) if tenant else None,
                'name': tenant.name if tenant else 'All Entities',
                'slug': tenant.slug if tenant else None
            } if tenant else None
        }
        
        return Response(analytics_data)
        
    except Exception as e:
        logger.error(f"Error in analytics_dashboard: {str(e)}")
        return Response(
            {'error': str(e)}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


def _generate_similarity_explanation(config, similarity_score, existing_configs):
    """
    Generate AI-powered explanation for similarity-based recommendations
    """
    try:
        explanations = []
        
        # Base confidence explanation
        if similarity_score >= 0.8:
            explanations.append("Excellent compatibility match with high confidence")
        elif similarity_score >= 0.6:
            explanations.append("Good compatibility match with moderate confidence")
        elif similarity_score >= 0.4:
            explanations.append("Fair compatibility match with some uncertainty")
        else:
            explanations.append("Limited compatibility match with low confidence")
        
        # Method-specific explanation
        explanations.append("Based on similarity analysis with existing fitments")
        
        # Find most similar existing configurations
        similar_configs = []
        for existing_config in existing_configs:
            # Calculate similarity factors
            year_diff = abs(config['year'] - existing_config['year'])
            make_match = config['makeName'] == existing_config['makeName']
            model_match = config['modelName'] == existing_config['modelName']
            drive_match = config['driveTypeName'] == existing_config['driveTypeName']
            fuel_match = config['fuelTypeName'] == existing_config['fuelTypeName']
            
            similarity_factors = []
            if make_match:
                similarity_factors.append("same make")
            if model_match:
                similarity_factors.append("same model")
            if drive_match:
                similarity_factors.append("same drive type")
            if fuel_match:
                similarity_factors.append("same fuel type")
            if year_diff <= 2:
                similarity_factors.append("similar year range")
            
            if similarity_factors:
                similar_configs.append({
                    'config': existing_config,
                    'factors': similarity_factors,
                    'year_diff': year_diff
                })
        
        # Sort by similarity (fewer year difference and more matching factors)
        similar_configs.sort(key=lambda x: (x['year_diff'], -len(x['factors'])))
        
        if similar_configs:
            top_similar = similar_configs[0]
            factors_text = ", ".join(top_similar['factors'])
            explanations.append(f"Strong similarity to {top_similar['config']['year']} {top_similar['config']['makeName']} {top_similar['config']['modelName']} based on {factors_text}")
        
        # Vehicle-specific factors
        vehicle_factors = []
        
        # Check for common makes (higher confidence)
        common_makes = ["Toyota", "Honda", "Ford", "Chevrolet", "Nissan", "BMW", "Mercedes-Benz"]
        if config['makeName'] in common_makes:
            vehicle_factors.append("popular make with extensive compatibility data")
        
        # Check for recent years (higher confidence)
        if config['year'] >= 2015:
            vehicle_factors.append("recent model year with updated specifications")
        elif config['year'] >= 2010:
            vehicle_factors.append("moderate age with established compatibility")
        else:
            vehicle_factors.append("older model with limited compatibility data")
        
        # Check for common drive types
        if config['driveTypeName'] in ["FWD", "AWD"]:
            vehicle_factors.append("common drive type with high compatibility")
        
        # Check for common fuel types
        if config['fuelTypeName'] in ["Gas", "Gasoline"]:
            vehicle_factors.append("standard fuel type with broad compatibility")
        
        if vehicle_factors:
            explanations.append("Vehicle factors: " + ", ".join(vehicle_factors))
        
        # Confidence level explanation
        if similarity_score >= 0.8:
            explanations.append("High confidence recommendation - safe to apply")
        elif similarity_score >= 0.6:
            explanations.append("Good confidence - recommended with verification")
        elif similarity_score >= 0.4:
            explanations.append("Moderate confidence - review before applying")
        else:
            explanations.append("Low confidence - manual verification recommended")
        
        return ". ".join(explanations)
        
    except Exception as e:
        logger.error(f"Error generating similarity explanation: {str(e)}")
        return "AI-generated similarity analysis based on existing fitment patterns"


def _generate_base_vehicle_explanation(config, base_vehicle_id):
    """
    Generate AI-powered explanation for base vehicle-based recommendations
    """
    try:
        explanations = []
        
        # Base confidence explanation
        explanations.append("High confidence recommendation based on base vehicle compatibility")
        explanations.append("Based on base vehicle relationship analysis")
        
        # Base vehicle relationship explanation
        explanations.append(f"Shares base vehicle ID {base_vehicle_id} with confirmed compatible vehicles")
        
        # Vehicle-specific factors
        vehicle_factors = []
        
        # Check for common makes
        common_makes = ["Toyota", "Honda", "Ford", "Chevrolet", "Nissan", "BMW", "Mercedes-Benz"]
        if config['makeName'] in common_makes:
            vehicle_factors.append("popular make with extensive compatibility data")
        
        # Check for recent years
        if config['year'] >= 2015:
            vehicle_factors.append("recent model year with updated specifications")
        elif config['year'] >= 2010:
            vehicle_factors.append("moderate age with established compatibility")
        else:
            vehicle_factors.append("older model with limited compatibility data")
        
        # Check for common drive types
        if config['driveTypeName'] in ["FWD", "AWD"]:
            vehicle_factors.append("common drive type with high compatibility")
        
        # Check for common fuel types
        if config['fuelTypeName'] in ["Gas", "Gasoline"]:
            vehicle_factors.append("standard fuel type with broad compatibility")
        
        if vehicle_factors:
            explanations.append("Vehicle factors: " + ", ".join(vehicle_factors))
        
        # Base vehicle method specific explanation
        explanations.append("Base vehicle method provides high confidence due to shared platform architecture")
        explanations.append("High confidence recommendation - safe to apply")
        
        return ". ".join(explanations)
        
    except Exception as e:
        logger.error(f"Error generating base vehicle explanation: {str(e)}")
        return "AI-generated base vehicle compatibility analysis"


@api_view(['POST'])
def bulk_update_status(request):
    """Bulk update fitment status"""
    try:
        data = request.data
        fitment_hashes = data.get('fitment_hashes', [])
        new_status = data.get('status', 'Active')
        
        if not fitment_hashes:
            return Response(
                {'error': 'No fitment hashes provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get tenant from request
        tenant_id = get_tenant_id_from_request(request)
        if not tenant_id:
            return Response(
                {'error': 'Tenant not found'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Filter fitments by tenant and provided hashes
        fitments = Fitment.objects.filter(
            hash__in=fitment_hashes,
            tenant_id=tenant_id
        )
        
        if not fitments.exists():
            return Response(
                {'error': 'No fitments found for the provided hashes'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Update status
        updated_count = fitments.update(itemStatus=new_status)
        
        return Response({
            'message': f'Successfully updated {updated_count} fitments to {new_status}',
            'updated_count': updated_count,
            'status': new_status
        })
        
    except Exception as e:
        logger.error(f"Error in bulk update status: {str(e)}")
        return Response(
            {'error': 'Failed to update fitment status'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@csrf_exempt
def approve_fitments(request):
    """Approve selected fitments (change status from readyToApprove to Active)"""
    try:
        data = request.data
        fitment_hashes = data.get('fitment_hashes', [])
        
        if not fitment_hashes:
            return Response(
                {'error': 'No fitment hashes provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get tenant from request
        tenant_id = get_tenant_id_from_request(request)
        if not tenant_id:
            return Response(
                {'error': 'Tenant not found'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Filter fitments by tenant and provided hashes
        # Handle both 'readyToApprove' and 'ReadyToApprove' status values
        fitments = Fitment.objects.filter(
            hash__in=fitment_hashes,
            tenant_id=tenant_id,
            itemStatus__in=['readyToApprove', 'ReadyToApprove']
        )
        
        if not fitments.exists():
            return Response(
                {'error': 'No fitments found for approval'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Update status to Active
        updated_count = fitments.update(itemStatus='Active')
        
        return Response({
            'message': f'Successfully approved {updated_count} fitments',
            'approved_count': updated_count
        })
        
    except Exception as e:
        logger.error(f"Error approving fitments: {str(e)}")
        return Response(
            {'error': 'Failed to approve fitments'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@csrf_exempt
def bulk_delete_fitments(request):
    """Bulk delete fitments"""
    logger.info(f"BULK DELETE CALLED: {request.method} {request.path}")
    try:
        data = request.data
        fitment_hashes = data.get('fitment_hashes', [])
        
        if not fitment_hashes:
            return Response(
                {'error': 'No fitment hashes provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get tenant from request
        tenant_id = get_tenant_id_from_request(request)
        
        if not tenant_id:
            return Response(
                {'error': 'Tenant not found'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Filter fitments by tenant and provided hashes
        fitments = Fitment.objects.filter(
            hash__in=fitment_hashes,
            tenant_id=tenant_id
        )
        
        if not fitments.exists():
            return Response(
                {'error': 'No fitments found to delete'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Soft delete fitments
        deleted_count = 0
        for fitment in fitments:
            fitment.soft_delete(deleted_by='bulk_delete_user')
            deleted_count += 1
        
        return Response({
            'message': f'Successfully deleted {deleted_count} fitments',
            'deleted_count': deleted_count
        })
        
    except Exception as e:
        logger.error(f"Error bulk deleting fitments: {str(e)}")
        return Response(
            {'error': 'Failed to delete fitments'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )


@api_view(['POST'])
@csrf_exempt
def reject_fitments(request):
    """Reject selected fitments (delete them)"""
    try:
        data = request.data
        fitment_hashes = data.get('fitment_hashes', [])
        
        if not fitment_hashes:
            return Response(
                {'error': 'No fitment hashes provided'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get tenant from request
        tenant_id = get_tenant_id_from_request(request)
        if not tenant_id:
            return Response(
                {'error': 'Tenant not found'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Filter fitments by tenant and provided hashes
        # Handle both 'readyToApprove' and 'ReadyToApprove' status values
        fitments = Fitment.objects.filter(
            hash__in=fitment_hashes,
            tenant_id=tenant_id,
            itemStatus__in=['readyToApprove', 'ReadyToApprove']
        )
        
        if not fitments.exists():
            return Response(
                {'error': 'No fitments found for rejection'}, 
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Delete fitments
        deleted_count = fitments.count()
        fitments.delete()
        
        return Response({
            'message': f'Successfully rejected and deleted {deleted_count} fitments',
            'rejected_count': deleted_count
        })
        
    except Exception as e:
        logger.error(f"Error rejecting fitments: {str(e)}")
        return Response(
            {'error': 'Failed to reject fitments'}, 
            status=status.HTTP_500_INTERNAL_SERVER_ERROR
        )
